#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ポケカ利益分析の結果を 1つの .xlsx（複数シート）にまとめる。
profit_analysis.py の計算ロジックを再利用して数値の一貫性を担保する。

シート構成:
  1. 利益商品一覧
  2. 30万円配分(方針A)
  3. 仕入れ先・販売先
  4. 仕入れ候補・相場ツール
出力: profit_analysis/output/ポケカ利益分析まとめ.xlsx
"""
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from profit_analysis import analyze, optimize  # noqa: E402

# ---- スタイル定義 -------------------------------------------------------
NAVY = "1A56DB"
LIGHT = "EEF3FF"
GREY = "F4F6FB"
HEAD_FONT = Font(bold=True, color="FFFFFF", size=11)
HEAD_FILL = PatternFill("solid", fgColor=NAVY)
TITLE_FONT = Font(bold=True, size=14, color=NAVY)
NOTE_FONT = Font(size=9, color="666666", italic=True)
BOLD = Font(bold=True)
LINK_FONT = Font(color="1A56DB", underline="single")
THIN = Side(style="thin", color="D0D7E2")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
YEN = '#,##0"円"'
PCT = '0.0"%"'


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEAD_FONT
        cell.fill = HEAD_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def widths(ws, ws_widths):
    for i, w in enumerate(ws_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def title(ws, text, span):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    c = ws.cell(row=1, column=1, value=text)
    c.font = TITLE_FONT
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24


def build():
    csv_path = os.path.join(HERE, "data", "pokemon_products.csv")
    budget = 300000
    cols = {"name": "商品名", "sale": "販売額", "fee": "手数料",
            "cost": "仕入れ値", "cap": "仕入れ可能数"}
    products, excluded = analyze(csv_path, budget, cols)

    wb = Workbook()

    # ====== Sheet1: 利益商品一覧 ======================================
    ws = wb.active
    ws.title = "利益商品一覧"
    title(ws, "■ 利益商品一覧（②③統合・利益額の大きい順 → 利益率の高い順）", 7)
    hdr = ["商品名", "仕入れ値", "販売額", "手数料", "1個利益", "利益率", "出典"]
    r0 = 3
    ws.append([])
    for j, h in enumerate(hdr, 1):
        ws.cell(row=r0, column=j, value=h)
    style_header(ws, r0, len(hdr))
    # 出典マップ
    import csv as _csv
    src = {}
    with open(csv_path, encoding="utf-8-sig") as f:
        for row in _csv.DictReader(f):
            src[row["商品名"]] = row.get("出典", "")
    r = r0 + 1
    for p in products:
        ws.cell(row=r, column=1, value=p["name"]).font = BOLD
        ws.cell(row=r, column=2, value=p["cost"]).number_format = YEN
        ws.cell(row=r, column=3, value=p["sale"]).number_format = YEN
        ws.cell(row=r, column=4, value=round(p["fee"], 1)).number_format = YEN
        ws.cell(row=r, column=5, value=round(p["profit"], 1)).number_format = YEN
        ws.cell(row=r, column=6, value=round(p["margin"] * 100, 1)).number_format = PCT
        ws.cell(row=r, column=7, value=src.get(p["name"], ""))
        for c in range(1, 8):
            ws.cell(row=r, column=c).border = BORDER
            if 2 <= c <= 6:
                ws.cell(row=r, column=c).alignment = RIGHT
        r += 1
    # 合計行
    ws.cell(row=r, column=1, value="合計/平均").font = BOLD
    ws.cell(row=r, column=2, value=f"=SUM(B{r0+1}:B{r-1})").number_format = YEN
    ws.cell(row=r, column=3, value=f"=SUM(C{r0+1}:C{r-1})").number_format = YEN
    ws.cell(row=r, column=4, value=f"=SUM(D{r0+1}:D{r-1})").number_format = YEN
    ws.cell(row=r, column=5, value=f"=SUM(E{r0+1}:E{r-1})").number_format = YEN
    ws.cell(row=r, column=6, value=f"=AVERAGE(F{r0+1}:F{r-1})").number_format = PCT
    for c in range(1, 8):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=LIGHT)
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).font = BOLD
    r += 2
    ws.cell(row=r, column=1,
            value="計算式: 1個利益 = 販売額 − 仕入れ値 − 手数料 ／ 利益率 = 1個利益 ÷ 仕入れ値").font = NOTE_FONT
    r += 1
    ws.cell(row=r, column=1,
            value="手数料の実態 = 販売額×10% + 3,300円（メルカリ10%＋送料相当と整合）").font = NOTE_FONT
    r += 1
    note = ("除外: " + "／".join(f"{n}:{why}" for n, why in excluded)) if excluded \
        else "除外した商品: なし（10商品すべて利益プラス）。原本テンプレ(空)は対象外。"
    ws.cell(row=r, column=1, value=note).font = NOTE_FONT
    widths(ws, [16, 12, 12, 12, 12, 9, 30])
    ws.freeze_panes = "A4"

    # ====== Sheet2: 30万円配分 ========================================
    ws2 = wb.create_sheet("30万円配分(方針A)")
    title(ws2, "■ 30万円投入時の最適配分（方針A=合計利益が最大）", 5)
    row = 3

    def scenario(ws, row, label, csv_file, note):
        prods, _ = analyze(csv_file, budget, cols)
        best_profit, used, counts = optimize(prods, budget)
        leftover = budget - used
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=12, color=NAVY)
        row += 1
        hdr = ["商品名", "仕入れ単価", "個数", "投入額", "見込み利益"]
        for j, h in enumerate(hdr, 1):
            ws.cell(row=row, column=j, value=h)
        style_header(ws, row, 5)
        row += 1
        idx_map = {p["name"]: p for p in prods}
        names_by_idx = {i: p for i, p in enumerate(prods)}
        tot_cost = tot_prof = 0
        for i in sorted(counts, key=lambda i: -names_by_idx[i]["profit"]):
            p = names_by_idx[i]
            n = counts[i]
            cs, pr = p["cost"] * n, p["profit"] * n
            tot_cost += cs
            tot_prof += pr
            ws.cell(row=row, column=1, value=p["name"]).font = BOLD
            ws.cell(row=row, column=2, value=p["cost"]).number_format = YEN
            ws.cell(row=row, column=3, value=n).alignment = CENTER
            ws.cell(row=row, column=4, value=cs).number_format = YEN
            ws.cell(row=row, column=5, value=round(pr, 1)).number_format = YEN
            for c in range(1, 6):
                ws.cell(row=row, column=c).border = BORDER
            row += 1
        ws.cell(row=row, column=1, value="合計").font = BOLD
        ws.cell(row=row, column=4, value=tot_cost).number_format = YEN
        ws.cell(row=row, column=5, value=round(tot_prof, 1)).number_format = YEN
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor=LIGHT)
            ws.cell(row=row, column=c).font = BOLD
            ws.cell(row=row, column=c).border = BORDER
        row += 1
        roi = tot_prof / tot_cost * 100 if tot_cost else 0
        ws.cell(row=row, column=1,
                value=f"投入額 {tot_cost:,.0f}円 ／ 端数(未使用) {leftover:,.0f}円 ／ "
                      f"見込み利益 {tot_prof:,.0f}円 ／ ROI {roi:.1f}%").font = BOLD
        row += 1
        ws.cell(row=row, column=1, value=note).font = NOTE_FONT
        return row + 2

    row = scenario(ws2, row, "【A-1】上限なし（タスク既定: 予算が許す限り仕入れ可）",
                   os.path.join(HERE, "data", "pokemon_products.csv"),
                   "※同一カードを大量に揃えられる前提。利益率最大のルリナに集中。端数4,000円は最安商品未満で使用不可。")
    row = scenario(ws2, row, "【A-2】各商品1枚まで（現実シナリオ: 同一カードを揃えにくい場合）",
                   os.path.join(HERE, "data", "pokemon_products_cap1.csv"),
                   "※在庫を各1枚に制限した0/1ナップサック。端数4,479円は使用不可。現実の物販に近い見込み。")
    widths(ws2, [30, 14, 8, 14, 16])

    # ====== Sheet3: 仕入れ先・販売先 ==================================
    ws3 = wb.create_sheet("仕入れ先・販売先")
    title(ws3, "■ 仕入れ先・販売先（2026年・手数料/特徴）", 4)
    # 販売先
    r = 3
    ws3.cell(row=r, column=1, value="● 販売先（高く売る＝手取りを残す）").font = Font(bold=True, size=12, color=NAVY)
    r += 1
    for j, h in enumerate(["販売先", "販売手数料", "特徴", "向き"], 1):
        ws3.cell(row=r, column=j, value=h)
    style_header(ws3, r, 4)
    sell = [
        ("magi（マギ）", "3%(+振込200円)", "トレカ特化・本物保証。手取り最大", "高額・コレクター品◎"),
        ("スニダン", "4.4〜6.05%", "プロ鑑定付きで買い手が安心", "高額品・鑑定需要"),
        ("Yahoo!フリマ", "5%", "メルカリより安く集客もそこそこ", "中価格帯"),
        ("ラクマ", "4.5〜10%(変動)", "楽天経済圏で有利", "楽天ユーザー"),
        ("メルカリ", "10%", "集客No.1＝最も売れやすい/相場が立つ", "回転重視・相場薄い品"),
        ("ヤフオク", "10%(プレ会員8.64%)", "競りで相場超えを狙える", "激レア・希少カード"),
    ]
    r += 1
    for row_data in sell:
        for j, v in enumerate(row_data, 1):
            cell = ws3.cell(row=r, column=j, value=v)
            cell.border = BORDER
            cell.alignment = LEFT
        ws3.cell(row=r, column=1).font = BOLD
        r += 1
    ws3.cell(row=r, column=1,
             value="おすすめ: 高額品はmagi(3%)、回転重視はメルカリ、激レアはヤフオク。"
                   "メルカリ(10%)→magi(3%)で販売額の約7%が利益増（10商品で約+89,000円）。").font = NOTE_FONT
    r += 3
    # 仕入れ先
    ws3.cell(row=r, column=1, value="● 仕入れ先（安く仕入れる＝原価を下げる）").font = Font(bold=True, size=12, color=NAVY)
    r += 1
    for j, h in enumerate(["仕入れ先", "狙い目", "URL", ""], 1):
        ws3.cell(row=r, column=j, value=h)
    style_header(ws3, r, 4)
    buy = [
        ("駿河屋（通販/あんしん買取）", "シングルが安め・在庫処分。宅配で手堅い", "https://www.suruga-ya.jp/"),
        ("カードショップのシングル/ワゴン", "入荷直後・店舗間の価格差・査定甘い掘出物", ""),
        ("フリマの相場割れ出品", "メルカリ/Yahoo!フリマの新着を即取り", ""),
        ("エコオク/エコリング", "手数料が安く未流通の「ウブ荷」", ""),
        ("コンビニ・量販店（新弾）", "定価でBOX確保→高騰待ち(セブン/ローソン/ファミマ等)", ""),
    ]
    r += 1
    for nm, aim, url in buy:
        ws3.cell(row=r, column=1, value=nm).font = BOLD
        ws3.cell(row=r, column=2, value=aim).alignment = LEFT
        if url:
            c = ws3.cell(row=r, column=3, value=url)
            c.hyperlink = url
            c.font = LINK_FONT
        for c in range(1, 4):
            ws3.cell(row=r, column=c).border = BORDER
        r += 1
    ws3.cell(row=r, column=1,
             value="⚠️ オリパ(パック販売)は期待値マイナス多く仕入れ向き×／BOXは再シュリンク詐欺に注意。").font = NOTE_FONT
    widths(ws3, [30, 40, 34, 18])

    # ====== Sheet4: 仕入れ候補・相場ツール ============================
    ws4 = wb.create_sheet("仕入れ候補・相場ツール")
    title(ws4, "■ 仕入れ候補・相場ツール（2026年5月 / 投資判断は自己責任）", 4)
    r = 3
    ws4.cell(row=r, column=1, value="● 今「歴史的に安い」銘柄（買取価格ベース。仕入れはこれより少し上）").font = Font(bold=True, size=12, color=NAVY)
    r += 1
    for j, h in enumerate(["銘柄", "現状", "コメント", "確認URL"], 1):
        ws4.cell(row=r, column=j, value=h)
    style_header(ws4, r, 4)
    cands = [
        ("セレナ SR(白熱のアルカナ/s11a)", "買取約8,000円(初動37,000円/PSA10≈25,000円)",
         "リストの「セレナ」に該当の可能性大。底値圏", "https://card-compass.jp/serena-sr/"),
        ("ルリナ SR(全5種)", "種類別に相場差", "リストの「ルリナ」。安い版を狙う",
         "https://premium.gamepedia.jp/pokeca/card/%E3%83%AB%E3%83%AA%E3%83%8A"),
        ("ソニア(全3種)", "種類別にばらつき", "リストの「ソニア」。安い版を拾う",
         "https://premium.gamepedia.jp/pokeca/card/%E3%82%BD%E3%83%8B%E3%82%A2"),
        ("メガリザードンY ex MUR", "944,000円(30日で−336,000円)", "高額帯の値ごろ例(30万では手が出ない)",
         "https://pokeka-atari.jp/ranking/crash"),
    ]
    r += 1
    for nm, st, cm, url in cands:
        ws4.cell(row=r, column=1, value=nm).font = BOLD
        ws4.cell(row=r, column=2, value=st).alignment = LEFT
        ws4.cell(row=r, column=3, value=cm).alignment = LEFT
        c = ws4.cell(row=r, column=4, value=url)
        c.hyperlink = url
        c.font = LINK_FONT
        for c in range(1, 5):
            ws4.cell(row=r, column=c).border = BORDER
        r += 1
    r += 1
    ws4.cell(row=r, column=1, value="● 毎日“今安いもの”を拾うための相場ツール").font = Font(bold=True, size=12, color=NAVY)
    r += 1
    for j, h in enumerate(["ツール", "用途", "URL", ""], 1):
        ws4.cell(row=r, column=j, value=h)
    style_header(ws4, r, 4)
    tools = [
        ("みんなのポケカ相場", "横断相場の確認", "https://pokeca-chart.com/"),
        ("暴落ランキング", "安く買える候補(毎日更新)", "https://pokeka-atari.jp/ranking/crash"),
        ("高騰ランキング", "高く売れる出口の確認", "https://altema.jp/pokemoncard/koutou"),
        ("トレチャ", "カードショップ横断・価格比較", "https://torechart.com/pokemon/priceup"),
        ("スニダン 買い時/売り時分析", "売買タイミング", "https://snkrdunk.com/articles/30531/"),
        ("駿河屋 買取検索", "仕入れ原価の目安(公開価格)", "https://www.suruga-ya.jp/kaitori/search_buy?category=501080033"),
    ]
    r += 1
    for nm, use, url in tools:
        ws4.cell(row=r, column=1, value=nm).font = BOLD
        ws4.cell(row=r, column=2, value=use).alignment = LEFT
        c = ws4.cell(row=r, column=3, value=url)
        c.hyperlink = url
        c.font = LINK_FONT
        for c in range(1, 4):
            ws4.cell(row=r, column=c).border = BORDER
        r += 1
    r += 1
    ws4.cell(row=r, column=1,
             value="型: ①暴落ランキングで底値の人気SAR/SR→②pokeca-chartで相場確認→"
                   "③メルカリ/Yahoo!フリマの相場割れを即取り→④出口はmagi(3%)。").font = NOTE_FONT
    r += 1
    ws4.cell(row=r, column=1,
             value="※価格は2026年5月時点の検索値。フリマの実出品最安は変動するため要現地確認。"
                   "値下がり中＝さらに下がるリスクあり。投資は余剰資金で。").font = NOTE_FONT
    widths(ws4, [30, 34, 44, 10])

    out_dir = os.path.join(HERE, "output")
    os.makedirs(out_dir, exist_ok=True)
    out = os.path.join(out_dir, "ポケカ利益分析まとめ.xlsx")
    wb.save(out)
    print("saved:", out)


if __name__ == "__main__":
    build()
