# -*- coding: utf-8 -*-
"""MT5『取引履歴レポート』(日本語・xlsx)のパーサ。ポジション一覧/約定一覧/集計を列位置で読む。"""
import openpyxl, re, datetime as dt
import pandas as pd

def _rows(path):
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    return list(wb[wb.sheetnames[0]].iter_rows(values_only=True))

def _sections(rows):
    idx = {}
    for i, r in enumerate(rows):
        s = "".join(str(v) for v in r if v is not None).strip()
        if s in ("ポジション一覧", "注文一覧", "注文", "約定一覧", "ポジション", "結果", "集計"):
            idx[s] = i
    return idx

def _ts(v):
    if v is None or v == "": return None
    if isinstance(v, dt.datetime): return pd.Timestamp(v)
    try: return pd.Timestamp(dt.datetime.strptime(str(v).strip(), "%Y.%m.%d %H:%M:%S"))
    except ValueError: return None

def _num(v):
    if v is None or v == "": return 0.0
    try: return float(str(v).replace(" ", "").replace(",", ""))
    except: return 0.0

def parse(path):
    rows = _rows(path); sec = _sections(rows)
    meta = {}
    for r in rows[:6]:
        if r and r[0] and str(r[0]).endswith(":"):
            meta[str(r[0]).rstrip(":")] = str(r[1]) if len(r) > 1 else ""
    # --- ポジション一覧 ---
    p0 = sec["ポジション一覧"] + 1
    hdr = [str(v) if v is not None else "" for v in rows[p0]]
    pos = []
    for r in (rows[p0 + 1:] if not hdr[0].startswith("銘柄") else []):
        if r is None or r[0] is None or str(r[0]).strip() == "" or _ts(r[0]) is None: break
        r = list(r) + [None] * (14 - len(r))
        pos.append(dict(open_time=_ts(r[0]), ticket=str(r[1]), symbol=str(r[2]), type=str(r[3]),
                        volume=_num(r[4]), open_price=_num(r[5]), sl=_num(r[6]), tp=_num(r[7]),
                        close_time=_ts(r[8]), close_price=_num(r[9]), commission=_num(r[10]),
                        swap=_num(r[11]), profit=_num(r[12]), comment=str(r[13]) if r[13] is not None else ""))
    pos = pd.DataFrame(pos)
    if len(pos):
        pos["net"] = pos["profit"] + pos["commission"] + pos["swap"]
    # --- 約定一覧(入出金・balance行を拾う) ---
    deals = []
    if "約定一覧" in sec:
        d0 = sec["約定一覧"] + 1
        dh = [str(v) if v is not None else "" for v in rows[d0]]
        for r in rows[d0 + 1:]:
            if r is None or r[0] is None or str(r[0]).strip() == "" or _ts(r[0]) is None: break
            r = list(r) + [None] * (16 - len(r))
            deals.append(dict(time=_ts(r[0]), deal=str(r[1]), symbol=str(r[2]), type=str(r[3]),
                              direction=str(r[4]), volume=_num(r[5]), price=_num(r[6]), order=str(r[7]),
                              commission=_num(r[8]), fee=_num(r[9]), swap=_num(r[10]), profit=_num(r[11]),
                              balance=_num(r[12]), comment=str(r[13]) if r[13] is not None else ""))
        deals = pd.DataFrame(deals)
    # --- 注文一覧(EAコメントが載る) ---
    orders = []
    if "注文" in sec or "注文一覧" in sec:
        o0 = sec.get("注文", sec.get("注文一覧")) + 1
        for r in rows[o0 + 1:]:
            if r is None or r[0] is None or str(r[0]).strip() == "" or _ts(r[0]) is None: break
            r = list(r) + [None] * (12 - len(r))
            orders.append(dict(time=_ts(r[0]), order=str(r[1]), symbol=str(r[2]), type=str(r[3]),
                               volume=str(r[4]), price=_num(r[5]), state=str(r[9]) if r[9] is not None else "",
                               comment=str(r[11]) if r[11] is not None else (str(r[10]) if r[10] is not None else "")))
    orders = pd.DataFrame(orders)
    # --- 未決済ポジション(履歴レポート末尾『ポジション』 or トレード口座レポートの『ポジション一覧』(銘柄が先頭)) ---
    open_pos = []; balance = None
    cand = []
    if "ポジション" in sec: cand.append(sec["ポジション"])
    hdr_first = str(rows[sec["ポジション一覧"] + 1][0]) if "ポジション一覧" in sec else ""
    if hdr_first.startswith("銘柄"): cand.append(sec["ポジション一覧"])
    for p1 in cand:
        hdr2 = [str(v) if v is not None else "" for v in rows[p1 + 1]]
        sym_first = hdr2[0].startswith("銘柄")
        for r in rows[p1 + 2:]:
            if r is None or r[0] is None or str(r[0]).strip() == "": break
            r = list(r) + [None] * (14 - len(r))
            if sym_first:
                if _ts(r[2]) is None: break
                open_pos.append(dict(symbol=str(r[0]), ticket=str(r[1]), open_time=_ts(r[2]), type=str(r[3]), volume=_num(r[4]),
                                     open_price=_num(r[5]), sl=_num(r[6]), tp=_num(r[7]), market=_num(r[8]), swap=_num(r[9]),
                                     profit=_num(r[10]), comment=str(r[11]) if r[11] is not None else ""))
            else:
                if _ts(r[0]) is None: break
                open_pos.append(dict(open_time=_ts(r[0]), ticket=str(r[1]), symbol=str(r[2]), type=str(r[3]), volume=_num(r[4]),
                                     open_price=_num(r[5]), sl=_num(r[6]), tp=_num(r[7]), market=_num(r[8]), swap=_num(r[9]),
                                     profit=_num(r[11]) if r[11] is not None else _num(r[10]),
                                     comment=str(r[12]) if r[12] is not None else (str(r[11]) if isinstance(r[11], str) else "")))
    for r in rows:
        if r and r[0] is not None and str(r[0]).startswith("残高"):
            vals = [v for v in r[1:] if v is not None]
            if vals: balance = _num(vals[0])
            break
    open_pos = pd.DataFrame(open_pos)
    summary_extra = {"balance": balance}
    # --- 集計(結果) ---
    summary = {}
    for key in ("結果", "集計"):
        if key in sec:
            for r in rows[sec[key] + 1: sec[key] + 40]:
                if not r: continue
                cells = [c for c in r if c is not None]
                for i in range(0, len(cells) - 1, 2):
                    k = str(cells[i]).rstrip(":").strip(); v = cells[i + 1]
                    if k and not isinstance(v, str) or (isinstance(v, str) and re.match(r"^-?[\d ,.]+%?$", v.strip() or "x")):
                        summary[k] = v
    summary.update(summary_extra)
    return meta, pos, deals, summary, orders, open_pos

if __name__ == "__main__":
    import sys, glob
    for f in sorted(glob.glob(sys.argv[1] + "/*.xlsx")):
        meta, pos, deals, summ = parse(f)
        print(f"\n===== {meta.get('口座','?')} | {meta.get('名前','')} =====")
        print(f"  positions={len(pos)}  deals={len(deals)}  期間={pos.open_time.min().date() if len(pos) else '-'}..{pos.close_time.max().date() if len(pos) else '-'}")
        if len(pos): print("  銘柄:", pos.groupby("symbol")["net"].agg(["count","sum"]).round(0).to_dict("index"))
        if len(deals):
            bal = deals[deals["type"].str.lower().str.contains("balance|入金|出金|credit", regex=True)]
            print("  balance/入出金行:", bal[["time","type","profit","comment"]].head(6).to_string(index=False) if len(bal) else "-")
            print("  約定コメント種別:", deals["comment"].str.extract(r"^\[?([A-Za-z_]+)")[0].value_counts().head(8).to_dict())
        print("  集計:", {k: summ[k] for k in list(summ)[:14]})
